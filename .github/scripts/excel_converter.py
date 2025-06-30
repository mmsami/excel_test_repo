#!/usr/bin/env python3
"""
Complete Excel Tracker for GitHub Actions
- Extracts only table structure (useful XML files)
- Adds Git-native cell change tracking
- Professional function names and clean code structure
"""

import os
import sys
import shutil
import zipfile
import glob
import git
import re
import time
import hashlib
from pathlib import Path

def extract_excel_tables_only(excel_path):
    """Extract only table definitions from Excel file - clean selective extraction"""
    excel_dir = Path(excel_path).with_suffix('')
    
    # Create directory if it doesn't exist
    os.makedirs(excel_dir, exist_ok=True)
    
    # Extract Excel as zip to temporary directory
    temp_dir = Path(f"{excel_dir}_temp")
    os.makedirs(temp_dir, exist_ok=True)
    
    with zipfile.ZipFile(excel_path, 'r') as zip_ref:
        zip_ref.extractall(temp_dir)
    
    # Copy ONLY table definitions (the useful files)
    tables_dir = temp_dir / 'xl/tables'
    if tables_dir.exists():
        dst_tables = excel_dir / 'xl/tables'
        shutil.copytree(tables_dir, dst_tables, dirs_exist_ok=True)
        print(f"✅ Extracted table structure from {excel_path}")
    else:
        print(f"⚠️ No tables found in {excel_path}")
    
    # Clean up temp directory
    shutil.rmtree(temp_dir)
    
    print(f"Extracted table definitions from {excel_path} to {excel_dir}/")
    return excel_dir

def extract_cells_for_git_tracking(excel_path):
    """Extract cell data to simple text files for Git-native diff tracking"""
    try:
        import openpyxl
        
        print(f"📊 Extracting cells for Git tracking: {excel_path}")
        start_time = time.time()
        
        wb = openpyxl.load_workbook(excel_path, data_only=True, read_only=True)
        
        # Create cells directory
        excel_dir = Path(excel_path).with_suffix('')
        cells_dir = excel_dir / 'cells'
        os.makedirs(cells_dir, exist_ok=True)
        
        total_cells = 0
        worksheets_processed = []
        
        for ws_name in wb.sheetnames:
            # Skip excluded worksheets for performance
            if any(skip in ws_name.lower() for skip in ['rawdata', 'import', 'temp', 'archive', 'backup']):
                print(f"   ⏭️ Skipping {ws_name} (excluded)")
                continue
            
            ws = wb[ws_name]
            
            # Create clean filename for worksheet
            safe_name = re.sub(r'[^\w\s-]', '', ws_name).strip()
            safe_name = re.sub(r'[-\s]+', '_', safe_name)
            worksheet_file = cells_dir / f"{safe_name}.txt"
            
            cell_lines = []
            cell_count = 0
            
            # Extract all cells with data
            for row in ws.iter_rows(values_only=False):
                for cell in row:
                    if cell.value is not None:
                        cell_ref = cell.coordinate
                        cell_value = str(cell.value)[:500]  # Truncate very long values
                        cell_type = type(cell.value).__name__
                        
                        # Format: CELL_REF=TYPE:VALUE (Git-diff friendly)
                        line = f"{cell_ref}={cell_type}:{cell_value}"
                        cell_lines.append(line)
                        cell_count += 1
                        
                        # Progress for large sheets
                        if cell_count % 10000 == 0:
                            print(f"      📈 {ws_name}: processed {cell_count} cells...")
            
            # Sort for consistent Git diffs (A1, A2, A3, B1, B2...)
            def sort_key(line):
                cell_ref = line.split('=')[0]
                # Extract column letters and row numbers
                col_match = re.match(r'([A-Z]+)(\d+)', cell_ref)
                if col_match:
                    col_letters, row_num = col_match.groups()
                    # Convert column letters to number for sorting
                    col_num = 0
                    for char in col_letters:
                        col_num = col_num * 26 + (ord(char) - ord('A') + 1)
                    return (int(row_num), col_num)
                return (0, 0)
            
            cell_lines.sort(key=sort_key)
            
            # Write worksheet cells to file
            with open(worksheet_file, 'w', encoding='utf-8') as f:
                f.write(f"# Worksheet: {ws_name}\n")
                f.write(f"# Cells: {cell_count}\n")
                f.write(f"# Generated: {time.strftime('%Y-%m-%d %H:%M:%S')}\n")
                f.write(f"# Excel File: {Path(excel_path).name}\n")
                f.write("\n")
                
                for line in cell_lines:
                    f.write(line + "\n")
            
            total_cells += cell_count
            worksheets_processed.append((ws_name, cell_count))
            print(f"   ✅ {ws_name}: {cell_count} cells → {worksheet_file.name}")
        
        # Create summary file
        create_cell_summary(cells_dir, total_cells, excel_path, worksheets_processed)
        
        wb.close()
        
        elapsed = time.time() - start_time
        print(f"   🎯 Total: {total_cells} cells in {elapsed:.2f}s")
        
        return cells_dir, total_cells
        
    except ImportError:
        print("❌ openpyxl not available - install with: pip install openpyxl")
        return None, 0
    except Exception as e:
        print(f"❌ Error extracting cells: {e}")
        return None, 0

def create_cell_summary(cells_dir, total_cells, excel_path, worksheets_processed):
    """Create a summary file for quick overview"""
    summary_file = cells_dir / "_SUMMARY.txt"
    
    with open(summary_file, 'w') as f:
        f.write(f"Excel Cell Data Summary\n")
        f.write(f"=" * 50 + "\n")
        f.write(f"Excel File: {Path(excel_path).name}\n")
        f.write(f"Total Cells: {total_cells:,}\n")
        f.write(f"Extraction Time: {time.strftime('%Y-%m-%d %H:%M:%S')}\n")
        f.write(f"File Size: {os.path.getsize(excel_path) / (1024*1024):.1f} MB\n")
        f.write(f"\nWorksheets Processed:\n")
        
        for ws_name, cell_count in worksheets_processed:
            f.write(f"  - {ws_name}: {cell_count:,} cells\n")
        
        # Create checksum for quick change detection
        all_content = ""
        for worksheet_file in sorted(cells_dir.glob("*.txt")):
            if worksheet_file.name != "_SUMMARY.txt":
                with open(worksheet_file, 'r') as wf:
                    all_content += wf.read()
        
        checksum = hashlib.md5(all_content.encode()).hexdigest()[:16]
        f.write(f"\nData Checksum: {checksum}\n")
        f.write(f"\nTo see changes: git diff HEAD~1 {cells_dir.relative_to(Path.cwd())}/\n")

def format_table_xml_files(directory):
    """Format table XML files with complete DxfId removal and formula preservation"""
    try:
        xml_count = 0
        
        for root, _, files in os.walk(directory):
            for file in files:
                if file.endswith('.xml'):
                    file_path = os.path.join(root, file)
                    
                    # Skip large XML files
                    if os.path.getsize(file_path) > 1024 * 1024:  # 1MB limit
                        continue
                    
                    try:
                        with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
                            content = f.read()
                        
                        # STEP 1: Extract and preserve formulas BEFORE any processing
                        formula_placeholders = {}
                        placeholder_counter = 0
                        
                        def extract_formula(match):
                            nonlocal placeholder_counter
                            tag_name = match.group(1)
                            formula_content = match.group(2)
                            
                            # Create unique placeholder
                            placeholder = f"__FORMULA_PLACEHOLDER_{placeholder_counter}__"
                            placeholder_counter += 1
                            
                            # Store original formula content with ZERO modification
                            formula_placeholders[placeholder] = {
                                'tag': tag_name,
                                'content': formula_content  # EXACT content from Excel
                            }
                            
                            # Replace with placeholder
                            return f"<{tag_name}>{placeholder}</{tag_name}>"
                        
                        # Extract calculatedColumnFormula
                        content = re.sub(
                            r'<(calculatedColumnFormula)>(.*?)</\1>',
                            extract_formula,
                            content,
                            flags=re.DOTALL
                        )
                        
                        # Extract regular formula tags
                        content = re.sub(
                            r'<(formula)>(.*?)</\1>',
                            extract_formula,
                            content,
                            flags=re.DOTALL
                        )
                        
                        # STEP 2: Remove ALL DxfId attributes completely
                        dxf_patterns = [
                            r'\s+DxfId="[^"]*"',           # DxfId="123"
                            r'\s+headerRowDxfId="[^"]*"',  # headerRowDxfId="200" 
                            r'\s+headerRowDxfID="[^"]*"',  # Case variation
                            r'\s+dataDxfId="[^"]*"',       # dataDxfId="200"
                            r'\s+dataDxfID="[^"]*"',       # Case variation
                            r'\s+totalsRowDxfId="[^"]*"',  # totalsRowDxfId="200"
                            r'\s+totalsRowDxfID="[^"]*"',  # Case variation
                            r'\s+dxfId="[^"]*"',           # lowercase version
                        ]
                        
                        total_removed = 0
                        for pattern in dxf_patterns:
                            before_matches = re.findall(pattern, content)
                            content = re.sub(pattern, '', content)
                            after_matches = re.findall(pattern, content)
                            removed = len(before_matches) - len(after_matches)
                            total_removed += removed
                        
                        if total_removed > 0:
                            print(f"🗑️ Removed {total_removed} DxfId attributes from {os.path.basename(file_path)}")
                        
                        # STEP 3: Entity decoding
                        content = content.replace('&amp;', '&')
                        content = content.replace('&lt;', '<')
                        content = content.replace('&gt;', '>')
                        content = content.replace('&quot;', '"')
                        content = content.replace('&apos;', "'")
                        
                        # STEP 4: Add requested newlines
                        content = add_table_xml_newlines(content)
                        
                        # STEP 5: Basic XML formatting (but NOT for formula content)
                        lines = content.split('\n')
                        formatted_lines = []
                        indent_level = 0
                        
                        for line in lines:
                            stripped = line.strip()
                            
                            if not stripped:
                                continue
                            
                            # Check if this line contains a formula placeholder
                            has_formula_placeholder = any(placeholder in line for placeholder in formula_placeholders.keys())
                            
                            if has_formula_placeholder:
                                # Don't indent lines with formula placeholders
                                formatted_lines.append(stripped)
                            elif stripped.startswith('</'):
                                # Closing tags
                                indent_level = max(0, indent_level - 1)
                                formatted_lines.append('    ' * indent_level + stripped)
                            elif stripped.endswith('/>'):
                                # Self-closing tags
                                formatted_lines.append('    ' * indent_level + stripped)
                            elif stripped.startswith('<') and not stripped.startswith('<?'):
                                # Opening tags
                                formatted_lines.append('    ' * indent_level + stripped)
                                if not stripped.endswith('/>'):
                                    indent_level += 1
                            elif stripped.startswith('<?'):
                                # XML declaration
                                formatted_lines.append(stripped)
                            else:
                                # Regular content
                                formatted_lines.append('    ' * indent_level + stripped)
                        
                        content = '\n'.join(formatted_lines)
                        
                        # STEP 6: Restore formulas with ZERO modification
                        for placeholder, formula_data in formula_placeholders.items():
                            tag_name = formula_data['tag']
                            original_content = formula_data['content']
                            
                            # Put back the original formula with NO indentation changes
                            restored_formula = f"<{tag_name}>{original_content}</{tag_name}>"
                            content = content.replace(f"<{tag_name}>{placeholder}</{tag_name}>", restored_formula)
                        
                        with open(file_path, 'w', encoding='utf-8') as f:
                            f.write(content)
                        
                        xml_count += 1
                        
                        # Report on formula preservation
                        if formula_placeholders:
                            print(f"✅ Preserved {len(formula_placeholders)} formulas unchanged in {os.path.basename(file_path)}")
                        
                        # Special handling for table files
                        if 'table' in file_path.lower():
                            print(f"✅ Processed table XML: {os.path.basename(file_path)}")
                            
                    except Exception as e:
                        print(f"Error formatting {file_path}: {e}")
        
        print(f"✅ Formatted {xml_count} XML files with complete DxfId removal and formula preservation")
    except Exception as e:
        print(f"Warning: XML formatting failed - {e}")

def add_table_xml_newlines(content):
    """Add specific newlines for better table XML readability"""
    
    # Add newlines before <tableColumn> tags
    content = re.sub(
        r'([^\n])(<tableColumn)',
        r'\1\n\2',
        content
    )
    
    # Add newlines before other table elements if needed
    table_elements = ['tableColumns', 'autoFilter', 'sortState']
    for element in table_elements:
        content = re.sub(
            f'([^\\n])(<{element})',
            r'\1\n\2',
            content
        )
    
    return content

def get_changed_excel_files():
    """Get list of changed Excel files from the most recent commit"""
    repo = git.Repo('.')
    
    try:
        if len(repo.heads) > 0:
            diffs = repo.git.diff('HEAD~1', '--name-only').split('\n')
        else:
            diffs = repo.git.ls_files().split('\n')
        
        # Filter for Excel files only
        excel_files = [f for f in diffs if f.endswith(('.xlsx', '.xlsm'))]
        return excel_files
    except git.exc.GitCommandError:
        all_files = repo.git.ls_files().split('\n')
        return [f for f in all_files if f.endswith(('.xlsx', '.xlsm'))]

def cleanup_temporary_files():
    """Clean up previously generated temporary files"""
    print("🧹 CLEANING UP TEMPORARY FILES...")
    
    cleanup_patterns = [
        '**/*_fromXML.xlsx',
        '**/*_fromXML.xlsm', 
        '**/*_overwritten.xlsx',
        '**/*_overwritten.xlsm',
        '**/*_temp*',
        '**/* - Copy*.xlsx',
        '**/* - Copy*.xlsm'
    ]
    
    cleaned_count = 0
    for pattern in cleanup_patterns:
        for file_path in glob.glob(pattern, recursive=True):
            try:
                if os.path.exists(file_path):
                    os.remove(file_path)
                    print(f"🗑️ Removed: {file_path}")
                    cleaned_count += 1
            except Exception as e:
                print(f"⚠️ Could not remove {file_path}: {e}")
    
    # Clean up old extracted directories (but keep the ones we create)
    for item in Path('.').rglob('*'):
        if (item.is_dir() and 
            item.name.endswith(('.xlsx', '.xlsm')) and
            item.name != item.parent.name and
            not (item / 'cells').exists() and
            not (item / 'xl' / 'tables').exists()):
            try:
                shutil.rmtree(item)
                print(f"🗂️ Removed directory: {item}")
                cleaned_count += 1
            except Exception as e:
                print(f"⚠️ Could not remove directory {item}: {e}")
    
    print(f"✅ Cleanup completed: {cleaned_count} items removed")

def process_excel_files_complete():
    """Process Excel files with table extraction and cell tracking"""
    print("=== STARTING EXCEL PROCESSING WITH TABLE + CELL TRACKING ===")
    start_time = time.time()
    
    # Get changed files first (most important)
    changed_files = get_changed_excel_files()
    excel_files = [f for f in changed_files if f.endswith(('.xlsx', '.xlsm'))]
    
    # If no changed files, get all files but filter heavily
    if not excel_files:
        all_files = glob.glob('**/*.xlsx', recursive=True) + glob.glob('**/*.xlsm', recursive=True)
        
        # Filter out temporary and generated files
        filtered_files = []
        for file_path in all_files:
            # Skip generated files
            if ('_fromXML' in file_path or 
                '_overwritten' in file_path or 
                '_temp' in file_path or
                ' - Copy' in file_path):
                print(f"⏭️ Skipping generated file: {file_path}")
                continue
            
            # Skip files inside extracted directories
            if ('/' in file_path and 
                any(part.endswith(('.xlsx', '.xlsm')) for part in Path(file_path).parts[:-1])):
                print(f"⏭️ Skipping nested file: {file_path}")
                continue
            
            # Check file size
            if os.path.exists(file_path):
                file_size_mb = os.path.getsize(file_path) / (1024 * 1024)
                if file_size_mb > 15:  # 15MB limit
                    print(f"⏭️ Skipping large file: {file_path} ({file_size_mb:.1f}MB)")
                    continue
                
                filtered_files.append((file_path, file_size_mb))
        
        # Sort by size and take only the smallest files for performance
        filtered_files.sort(key=lambda x: x[1])
        excel_files = [f[0] for f in filtered_files[:3]]
        
        print(f"📊 Filtered from {len(all_files)} to {len(excel_files)} files")
        for file_path, size_mb in filtered_files[:3]:
            print(f"   ✅ Will process: {file_path} ({size_mb:.1f}MB)")
    
    processed_count = 0
    
    for excel_file in excel_files:
        # Double-check filters
        if ('_fromXML' in excel_file or 
            '_overwritten' in excel_file or
            not os.path.exists(excel_file)):
            continue
        
        # Timeout check
        elapsed = time.time() - start_time
        if elapsed > 120:  # 2 minutes max
            print(f"⏱️ TIMEOUT: Stopping after {elapsed:.1f}s")
            break
        
        file_size_mb = os.path.getsize(excel_file) / (1024 * 1024)
        print(f"🔄 Processing {excel_file} ({file_size_mb:.1f}MB)")
        
        file_start = time.time()
        
        try:
            # STEP 1: Extract table definitions only
            extract_dir = extract_excel_tables_only(excel_file)
            if not extract_dir:
                continue
            
            # STEP 2: Format table XML files
            print("🎨 Formatting table XML files...")
            format_table_xml_files(extract_dir)
            
            # STEP 3: Extract cells for Git-native diff tracking
            print("📊 Extracting cells for Git-native tracking...")
            cells_dir, cell_count = extract_cells_for_git_tracking(excel_file)
            if cells_dir:
                print(f"✅ Extracted {cell_count:,} cells to {cells_dir}")
            
            # Add everything to git
            repo = git.Repo('.')
            repo.git.add(str(extract_dir))
            
            # Create backup copy
            fromxml_path = Path(excel_file).with_name(
                f"{Path(excel_file).stem}_fromXML{Path(excel_file).suffix}"
            )
            shutil.copy2(excel_file, fromxml_path)
            repo.git.add(str(fromxml_path))
            
            processed_count += 1
            file_elapsed = time.time() - file_start
            total_elapsed = time.time() - start_time
            
            print(f"✅ Processed {excel_file} in {file_elapsed:.1f}s (total: {total_elapsed:.1f}s)")
            
        except Exception as e:
            print(f"❌ Error processing {excel_file}: {e}")
            continue
    
    total_elapsed = time.time() - start_time
    print(f"=== COMPLETED: {processed_count} files in {total_elapsed:.1f}s ===")
    return excel_files

def find_table_xml_directories():
    """Find directories with table XML changes"""
    changed_files = get_changed_excel_files()
    
    xml_dirs = set()
    for changed_file in changed_files:
        if changed_file.endswith('.xml') or '/xl/tables/' in changed_file:
            path = Path(changed_file)
            current_dir = path.parent
            
            while str(current_dir) != '.':
                if (current_dir / 'xl' / 'tables').exists():
                    xml_dirs.add(str(current_dir))
                    break
                
                if current_dir == Path('.') or current_dir.parent == current_dir:
                    break
                
                current_dir = current_dir.parent
    
    print(f"Found {len(xml_dirs)} directories with table changes")
    return list(xml_dirs)

def generate_table_diff_report(xml_dirs):
    """Generate diff report for table structure changes"""
    try:
        reports_dir = Path("table-diff-reports")
        os.makedirs(reports_dir, exist_ok=True)
        repo = git.Repo('.')
        
        for xml_dir in xml_dirs:
            summary_file = reports_dir / f"{Path(xml_dir).name}-table-changes.md"
            
            with open(summary_file, 'w') as f:
                f.write(f"# Table Structure Changes for {xml_dir}\n\n")
                
                table_changes = []
                total_changes = 0
                
                tables_path = os.path.join(xml_dir, 'xl', 'tables')
                if os.path.exists(tables_path):
                    for root, _, files in os.walk(tables_path):
                        for file in files:
                            if file.endswith('.xml'):
                                file_path = os.path.join(root, file)
                                rel_path = os.path.relpath(file_path, '.')
                                
                                try:
                                    diff = repo.git.diff('HEAD~1', file_path)
                                    
                                    if not diff:
                                        continue
                                        
                                    added = diff.count('\n+')
                                    removed = diff.count('\n-')
                                    changes = added + removed
                                    
                                    if changes > 0:
                                        total_changes += changes
                                        table_changes.append((rel_path, changes))
                                        
                                except Exception as e:
                                    f.write(f"Error analyzing {rel_path}: {e}\n\n")
                
                f.write(f"**Total changes:** {total_changes} lines\n\n")
                
                if table_changes:
                    f.write("## Table Structure Changes\n\n")
                    for file_path, changes in table_changes:
                        f.write(f"- **{file_path}**: {changes} lines changed\n")
                    f.write("\n")
                else:
                    f.write("No table structure changes detected.\n\n")
            
            print(f"Created table diff report for {xml_dir}")
        
        return reports_dir
    except Exception as e:
        print(f"Warning: Failed to generate table diff report - {e}")
        return None

def generate_cell_changes_summary():
    """Generate a summary of cell changes using Git's native diff"""
    try:
        reports_dir = Path("cell-diff-reports")
        os.makedirs(reports_dir, exist_ok=True)
        
        repo = git.Repo('.')
        
        # Find all cell directories
        cell_dirs = []
        for item in Path('.').rglob('cells'):
            if item.is_dir():
                cell_dirs.append(item)
        
        if not cell_dirs:
            print("No cell tracking directories found")
            return
        
        summary_file = reports_dir / "cell_changes_summary.md"
        
        with open(summary_file, 'w') as f:
            f.write("# 📊 Excel Cell Changes Summary\n\n")
            f.write(f"Generated: {time.strftime('%Y-%m-%d %H:%M:%S')}\n\n")
            
            total_files_changed = 0
            
            for cells_dir in cell_dirs:
                excel_name = cells_dir.parent.name
                f.write(f"## 📁 {excel_name}\n\n")
                
                # Check for changes in this cells directory
                try:
                    diff_output = repo.git.diff('HEAD~1', str(cells_dir), name_only=True)
                    if diff_output:
                        changed_files = diff_output.strip().split('\n')
                        f.write(f"**Files with changes:** {len(changed_files)}\n\n")
                        
                        for changed_file in changed_files:
                            if changed_file.endswith('.txt'):
                                worksheet_name = Path(changed_file).stem
                                f.write(f"### 📋 {worksheet_name}\n")
                                
                                # Get the actual diff
                                try:
                                    diff_content = repo.git.diff('HEAD~1', changed_file)
                                    if diff_content:
                                        # Count changes
                                        added_lines = diff_content.count('\n+')
                                        removed_lines = diff_content.count('\n-')
                                        
                                        f.write(f"- **Changes:** +{added_lines} -{removed_lines} lines\n")
                                        f.write(f"- **View diff:** `git diff HEAD~1 {changed_file}`\n\n")
                                        
                                        total_files_changed += 1
                                except:
                                    f.write("- Could not get diff details\n\n")
                    else:
                        f.write("**No changes detected**\n\n")
                        
                except Exception as e:
                    f.write(f"**Error checking changes:** {e}\n\n")
            
            f.write(f"---\n**Total Excel files with cell changes:** {total_files_changed}\n\n")
            f.write("💡 **To see detailed changes:**\n")
            f.write("- Click on any `.txt` file in the cells/ directories\n")
            f.write("- View the file history in GitHub\n")
            f.write("- Git will show you exactly which cells changed!\n")
        
        print(f"📊 Generated cell changes summary: {summary_file}")
        
    except Exception as e:
        print(f"Warning: Could not generate cell diff summary - {e}")

def commit_all_changes():
    """Commit any changes and push to repository"""
    repo = git.Repo('.')
    
    if repo.is_dirty() or len(repo.untracked_files) > 0:
        try:
            token = os.environ.get('GITHUB_TOKEN')
            
            origin_url = repo.remotes.origin.url
            if origin_url.startswith('https://'):
                new_url = f"https://x-access-token:{token}@github.com/{'/'.join(origin_url.split('/')[3:])}"
                repo.remotes.origin.set_url(new_url)
            
            branch = os.environ.get('GITHUB_REF', 'refs/heads/main').replace('refs/heads/', '')
            
            try:
                repo.git.pull('--rebase', 'origin', branch)
            except git.exc.GitCommandError as e:
                print(f"Pull error (non-critical): {e}")
            
            repo.git.add(A=True)
            commit_msg = "Excel Analysis: Table Structure + Cell Data Tracking [skip ci]"
            repo.git.commit('-m', commit_msg)
            
            repo.git.push('origin', branch)
            print(f"Changes committed and pushed to {branch}")
            
        except git.exc.GitCommandError as e:
            print(f"Git push error: {e}")

def main():
    """Main function to process Excel files with table and cell tracking"""
    print("🚀 STARTING: Excel Table Structure + Cell Data Analysis...")
    
    # STEP 1: Clean up temporary files
    print("🧹 STEP 1: Starting cleanup...")
    cleanup_temporary_files()
    print("✅ STEP 1: Cleanup completed")
    
    # STEP 2: Process Excel files with table extraction and cell tracking
    print("📁 STEP 2: Starting Excel file processing...")
    excel_files = process_excel_files_complete()
    print(f"✅ STEP 2: Excel processing completed - {len(excel_files)} files")
    
    # STEP 3: Find directories with table changes
    print("🔍 STEP 3: Finding table XML directories...")
    xml_dirs = find_table_xml_directories()
    print(f"✅ STEP 3: Found {len(xml_dirs)} directories with changes")
    
    # STEP 4: Generate table structure diff reports
    print("📊 STEP 4: Generating table diff reports...")
    if xml_dirs:
        generate_table_diff_report(xml_dirs)
    print("✅ STEP 4: Table diff reports completed")
    
    # STEP 5: Generate cell changes summary
    print("📋 STEP 5: Generating cell changes summary...")
    generate_cell_changes_summary()
    print("✅ STEP 5: Cell changes summary completed")
    
    # STEP 6: Commit everything to Git
    print("📤 STEP 6: Committing changes...")
    commit_all_changes()
    print("✅ STEP 6: Changes committed and pushed")
    
    print("🏁 COMPLETED: Excel Analysis")
    print("=" * 60)
    print("📊 What was tracked:")
    print("  ✅ Table Structure Changes (xl/tables/*.xml)")
    print("  ✅ Cell Data Changes (cells/*.txt)")
    print("  ✅ Git-native diff tracking for all changes")
    print("  ✅ Human-readable reports generated")
    print("")
    print("🔍 To view changes:")
    print("  • GitHub: Click on any file → History → Select commit")
    print("  • Command line: git diff HEAD~1")
    print("  • Reports: Check table-diff-reports/ and cell-diff-reports/")
    print("")
    print("📁 Files extracted per Excel file:")
    print("  • xl/tables/ - Table structure definitions")
    print("  • cells/ - Cell data in Git-friendly format")
    print("  • No metadata files (Content_Types.xml, _rels/) - removed as requested")

if __name__ == "__main__":
    main()
